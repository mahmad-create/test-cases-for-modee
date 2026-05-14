"""
Build 5-column test case TSV + XLSX from extracted SRS text.
Source: Requirements/BRD/_extracted_SRS_Front_End_ACC_V1.4.txt (generated from DOCX).
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXTRACT = ROOT / "Requirements" / "BRD" / "_extracted_SRS_Front_End_ACC_V1.4.txt"
OUT_TSV = ROOT / "Test cases" / "TestCases_SRS-FrontEnd-ACC-V1.4.tsv"
OUT_XLSX = ROOT / "Test cases" / "TestCases_SRS-FrontEnd-ACC-V1.4.xlsx"

STORY = "SRS Front End - ACC – V1.4"
QA = "Muna Ahmad"

# Split after FE screen narratives; part 11 is mostly pasted tables — still one service + tables
SECTION_LABELS = [
    "§1 براءة ذمة (incl. shared portal shell)",
    "§2 كتاب التزام",
    "§3 تغيير كفيل حسم",
    "§4 الحصول على كشف حساب بحالة القرض/التمويل",
    "§5 الحصول على قرض",
    "§6 الحصول على تمويل",
    "§7 الضم والتوحيد",
    "§8 الإفراز",
    "§9 فك حجز",
    "§10 طلب كشف ميداني وصرف قسط",
    "§11 إصدار كتاب تسوية شيكات (+ trailing tables in extract)",
]


def templates_for_section(label: str, big: bool) -> list[str]:
    """Return Description strings for one section (Verify-first, per workspace QA rules)."""
    base = [
        f"Verify that when the user completes the standard submission path from the electronic services home (locate the service via icons or search, open the service card, accept the تعهد, start إبدأ الخدمة or the تقديم flow, complete mandatory fields per the SRS, and submit إرسال الطلب), the application displays a success message and reference number and lists the request in طلباتي with the expected columns. Ref: {label}.",
        f"Verify that using رجوع between steps does not discard a valid draft where applicable, and that التالي advances each step of multi-step flows in line with the SRS. Ref: {label}.",
        f"Verify that دليل المستخدم (PDF), شرح الخدمة (video), and فترة إتمام الخدمة open and load without error. Ref: {label}.",
        f"Verify that the إمكانية الوصول path is available and the core service flow can be completed. Ref: {label}.",
        f"Verify that the system blocks submission when تعهد is required but not accepted, and presents clear validation messaging. Ref: {label}.",
        f"Verify that invalid or incomplete mandatory fields prevent submission and the user receives actionable guidance. Ref: {label}.",
        f"Verify that طلباتي allows editing only for permitted statuses (for example مسودة or إرجاع) and shows an explanatory message when editing is not allowed. Ref: {label}.",
        f"Verify that delete or cancel in طلباتي is permitted only for allowed statuses, requires explicit confirmation for irreversible actions, and sets the status to ملغي when confirmed. Ref: {label}.",
        f"Verify that حالة الطلب displays lifecycle states with dates consistent with the SRS excerpt. Ref: {label}.",
        f"Verify that identity-related fields are auto-populated from Sand, MoDEE, or CSPD where the SRS states they should be, and that the UI handles external service unavailability gracefully. Ref: {label}.",
        f"Verify that the employee inbox lists incoming requests, that تحديث refreshes the list, and that عرض opens a request for review or return per the SRS. Ref: {label}.",
        f"Verify that the notification bell badge count matches the items shown in the notification list. Ref: {label}.",
    ]
    extra: list[str] = []
    if big:
        extra = [
            f"Verify that the attachments area supports نوع المرفق، وصف المرفق، and ملاحظات; allows upload, supports preview (عرض), and enforces remove or replace behavior per UI rules. Ref: {label}.",
            f"Verify that guarantor and employer response flows (موافقة or رفض, amounts, attachments, and mandatory rejection reasons) route and persist correctly per the SRS. Ref: {label}.",
            f"Verify that any fee or payment step required by the flow blocks progression until the payment condition is satisfied. Ref: {label}.",
            f"Verify that after a successful submit, unrelated historical requests and access to other services remain available and unchanged. Ref: {label}.",
            f"Verify that large multi-step forms (many التالي steps) remain responsive during entry and that a double-click on submit does not create duplicate requests. Ref: {label}.",
        ]
    return base + extra


def main() -> None:
    text = EXTRACT.read_text(encoding="utf-8")
    parts = text.split("\nانتهى\n")
    if len(parts) < 11:
        raise SystemExit(f"Unexpected split count: {len(parts)}")

    rows: list[tuple[str, str, str, str, str]] = []
    n = 0
    for idx, part in enumerate(parts[:11]):
        label = SECTION_LABELS[idx] if idx < len(SECTION_LABELS) else f"Section {idx}"
        big = idx in (4, 5)  # قرض، تمويل
        for desc in templates_for_section(label, big):
            n += 1
            rows.append((f"TC-SRSFE-{n:04d}", STORY, desc, "", QA))

    OUT_TSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_TSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        w.writerow(["Test Case ID", "Story Title", "Description", "Status", "QA Name"])
        w.writerows(rows)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="DDEEFF")

    ws.append(["Test Case ID", "Story Title", "Description", "Status", "QA Name"])
    for r in rows:
        ws.append(list(r))
    for r_i in range(1, ws.max_row + 1):
        for c_i in range(1, 6):
            cell = ws.cell(r_i, c_i)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_i == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                if r_i % 2 == 0:
                    cell.fill = alt_fill
    ws.freeze_panes = "A2"
    for col, w in {"A": 18, "B": 36, "C": 100, "D": 10, "E": 12}.items():
        ws.column_dimensions[col].width = w
    wb.save(OUT_XLSX)

    print(f"Wrote {len(rows)} rows to {OUT_TSV.name} and {OUT_XLSX.name}")


if __name__ == "__main__":
    main()
