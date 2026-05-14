"""
Generate BRD-style manual test cases from SRS Front End ACC V1.4 extract.
Writes: Test cases/TestCases_BRD_SRS_FrontEnd_ACC_V1_4_<stamp>.{md,tsv,xlsx}
  (underscore before 4 avoids pathlib treating ``.4_<stamp>`` as the file suffix.)

Environment (optional):
  BRD_STABLE_OUTPUT=1 — write ``TestCases_BRD_SRS_FrontEnd_ACC_V1_4_latest.*`` (for CI).
  BRD_EXPORT_STAMP=mytext — suffix instead of default date (ignored if BRD_STABLE_OUTPUT=1).
"""
from __future__ import annotations

import csv
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_STAMP = "20260513"
STORY = "SRS Front End - ACC – V1.4"
QA = "Muna Ahmad"


def _output_base() -> Path:
    if os.environ.get("BRD_STABLE_OUTPUT") == "1":
        return ROOT / "Test cases" / "TestCases_BRD_SRS_FrontEnd_ACC_V1_4_latest"
    stamp = os.environ.get("BRD_EXPORT_STAMP", _DEFAULT_STAMP).strip() or _DEFAULT_STAMP
    return ROOT / "Test cases" / f"TestCases_BRD_SRS_FrontEnd_ACC_V1_4_{stamp}"

# (description only — IDs assigned in order)
_DESCS: list[str] = [
    # A — Shared electronic services portal
    "Verify that the authenticated applicant reaches the electronic services home screen with service icons and can open the main home action when offered. Ref: shared portal — electronic services home.",
    "Verify that the applicant can locate a service using the labelled service icons and start the flow with تقديم as described in the SRS. Ref: shared portal — service discovery.",
    "Verify that an applicant who uses إمكانية الوصول can complete the offered accessibility path for submitting a request. Ref: shared portal — accessibility.",
    "Verify that opening the user menu shows both الملف الشخصي and تسجيل الخروج and that each option navigates or signs out as specified. Ref: shared portal — profile menu.",
    "Verify that the profile screen displays Sand-linked identity fields (full name, email, national ID, phone) and directs profile updates through Sand when the user edits data there. Ref: shared portal — Sand profile.",
    "Verify that when notifications exist, a red badge count appears on the bell and opening الجرس lists auto-generated notification descriptions. Ref: shared portal — notifications.",
    "Verify that طلباتي lists each request with رقم الطلب، إسم الخدمة، تاريخ تقديم الخدمة، حالة الطلب، نتيجة الطلب، and التاريخ المتوقع لإتمام الخدمة as populated by the system. Ref: shared portal — طلباتي grid.",
    "Verify that تعديل الطلب is available only when the request status is حفظ كمسودة or إرجاع, and is blocked with appropriate behaviour otherwise. Ref: shared portal — edit eligibility.",
    "Verify that حالة الطلب shows timeline rows for حفظ كمسودة، قيد التنفيذ، and ملغي each with its date, and that إغلاق returns to طلباتي. Ref: shared portal — status timeline.",
    "Verify that حذف الطلب for eligible statuses shows a non-reversible warning, that تأكيد الإلغاء sets status to ملغي with success messaging, and that الرجوع cancels without cancelling the request. Ref: shared portal — delete/cancel.",
    "Verify that search in طلباتي using إسم الطلب، رقم الطلب، حالة الطلب، and تاريخ الطلب returns the matching request row(s). Ref: shared portal — search.",
    "Verify that الأسئلة الأكثر تكراراً loads FAQ items, expands an answer via السهم, and supports search by إسم الطلب with تصنيف حسب. Ref: shared portal — FAQ.",
    "Verify that إتصل بنا displays contact channels (phone, email, hours, website, map, social) supplied by the system for ACC. Ref: shared portal — contact.",
    # B — براءة ذمة / إيقاف اقتطاع
    "Verify that on the service card for براءة ذمة / إيقاف اقتطاع the applicant sees auto-filled summary fields (وصف الخدمة، متلقي الخدمة، رسوم الخدمة، المستندات المطلوبة، شروط الخدمة) before accepting التعهد. Ref: §1 براءة ذمة — service card.",
    "Verify that the applicant cannot start إبدأ الخدمة until the التعهد checkbox is selected, and that رجوع returns to the electronic services home. Ref: §1 براءة ذمة — تعهد gate.",
    "Verify that for نوع الكتاب إيقاف اقتطاع the applicant completes auto-filled identity fields from CSPD/Sand/MoDEE plus dropdowns (نوع الكتاب، الجهة الطالبة name/number، رقم القرض) and that نوع الجهة الطالبة is auto-derived, then إرسال الطلب shows success with رقم الطلب and موافق returns to طلباتي. Ref: §1 — إيقاف اقتطاع path.",
    "Verify that for نوع الكتاب براءة ذمة with requesting party type أخرى or بنك or جهة عمل the applicant completes the branch-specific dropdowns and submit shows success with reference number. Ref: §1 — براءة ذمة (جهة أخرى/بنك/عمل).",
    "Verify that for نوع الكتاب براءة ذمة with نوع الجهة الطالبة لمن يهمه الأمر the applicant completes the reduced dropdown set (including رقم القرض) and successful submit returns to طلباتي via موافق. Ref: §1 — براءة ذمة (لمن يهمه الأمر).",
    "Verify that the ACC employee inbox lists incoming requests, تحديث refreshes the queue, and عرض opens the request for review or return to the applicant per SRS. Ref: §1 — Back End employee flow.",
    # C — كتاب التزام
    "Verify that for كتاب إلتزام with الكتاب يشمل تمويلات/إدانات and requesting party أخرى/بنك/عمل the applicant can select الكتاب يشمل، سبب طلب الكتاب، نوع الجهة الطالبة، الجهة الطالبة, optionally indicate إدراج بيانات الأموال غير المنقولة, enter ملاحظات, and submit successfully. Ref: §2 كتاب التزام — branch (تمويلات/إدانات × جهة).",
    "Verify that the same service variant for لمن يهمه الأمر uses the correct dropdown subset (without الجهة الطالبة name field where omitted) and still supports the إدراج أموال غير منقولة flag and ملاحظات. Ref: §2 — لمن يهمه الأمر.",
    "Verify that for الكتاب يشمل قرض محدد with أخرى/بنك/عمل the applicant supplies رقم القرض plus سبب طلب الكتاب and completes submit with confirmation and رقم الطلب. Ref: §2 — قرض محدد (جهة).",
    "Verify that for قرض محدد with لمن يهمه الأمر the applicant completes رقم القرض and related dropdowns and receives the standard success dialog. Ref: §2 — قرض محدد (لمن يهمه الأمر).",
    # D — تغيير كفيل حسم
    "Verify that after إبدأ الخدمة the applicant selects رقم القرض, advances with التالي, sees old حسم summary (إسم الكفيل، قيمة القرض، قيمة الإقتطاع، جهة العمل، رقم القرض), selects one حسم via إختيار, and reaches the new guarantor entry screen. Ref: §3 تغيير كفيل حسم — applicant part 1.",
    "Verify that إضافة for the new guarantor persists required manual fields and auto-filled name from national ID inputs, shows جهة العمل from social security or military/civilian retirement when applicable, and saves on إضافة. Ref: §3 — new guarantor modal.",
    "Verify that تعديل and حذف on the new guarantor row behave per SRS (save changes, delete removes row) and empty-closing actions return to the main new-guarantor list. Ref: §3 — edit/delete new guarantor.",
    "Verify that the attachments step supports نوع المرفق، وصف المرفق، ملاحظات, إضافة ملاحظة, عرض, and تحميل before إرسال الطلب. Ref: §3 — attachments.",
    "Verify that after submit the new guarantor receives notification, opens the request via تعديل الطلب from طلباتي, answers نعم/لا with notes, and إرسال routes to ACC staff on approve or back to applicant on reject. Ref: §3 — new guarantor response.",
    "Verify that on employer approval path the guarantor can print الكتاب and رد جهة العمل, enter القيمة الموافق عليها and attach موافقة جهة العمل when موافق, or enter سبب الرفض when غير موافق, and إرسال completes the branch. Ref: §3 — employer response (موافق / غير موافق).",
    # E — كشف حساب
    "Verify that after التعهد and إبدأ الخدمة the applicant completes رقم القرض and سبب طلب الكتاب from dropdowns, enters ملاحظات, submits with إرسال الطلب, and receives success with رقم الطلب. Ref: §4 الحصول على كشف حساب بحالة القرض/التمويل.",
    # F — الحصول على قرض (representative)
    "Verify that the loan request flow exposes دليل المستخدم (PDF)، شرح الخدمة (video)، and فترة إتمام الخدمة from the service introduction screen. Ref: §5 الحصول على قرض — aids.",
    "Verify that multi-step navigation uses التالي and رجوع without losing valid entries where the SRS implies draft retention between steps. Ref: §5 — navigation.",
    # G — الحصول على تمويل (deeper)
    "Verify that the financing applicant completes مكان القيد، المؤهل العلمي، طريقة التعرف على المؤسسة plus structured address (المحافظة، المنطقة، اللواء، وصف العنوان) with auto email/phone. Ref: §6 الحصول على تمويل — identity & address.",
    "Verify that reference (معرف) fields and insurance health indicators enable وصف المرض when any condition checkbox is selected, per SRS validation. Ref: §6 — reference & health disclosure.",
    "Verify that adding a sub-goal with طريقة الصرف شركة loads company registry fields from دائرة مراقبة الشركات after رقم المنشأة and تاريخ التسجيل, and saves on إضافة. Ref: §6 — غاية فرعية (شركة).",
    "Verify that طريقة الصرف متعهد أو فرد populates الإسم from CSPD after national ID triplet and saves the sub-goal row. Ref: §6 — غاية فرعية (فرد).",
    "Verify that project section recalculates قيمة القسط المتوقع when مدة السداد بالسنوات changes and shows رقم نقابة المهندسين الزراعيين when نوع المشروع is مشروع المهندسين الزراعيين. Ref: §6 — project & repayment.",
    # H — الضم والتوحيد
    "Verify that the applicant selects الأرض المطلوب ضمها, uses التالي, and adds a parcel owned by self with the ملك لمقدم الطلب path and إضافة saves parcel summary. Ref: §7 الضم والتوحيد — self-owned parcel.",
    "Verify that adding a parcel not owned by self captures owner national ID, طريقة الموافقة على ضم الأرض, and land location fields from دائرة الأراضي والمساحة integration. Ref: §7 — third-party owner parcel.",
    "Verify that تحميل آلي من الجهة accepts الجهة plus رقم الكتاب or الرقم المميز and loads the authority document into attachments. Ref: §7 — automatic authority load.",
    "Verify that after submit the guarantor is notified and can approve or reject with notes from طلباتي per the ضم والتوحيد narrative. Ref: §7 — guarantor gate.",
    # I — الإفراز
    "Verify that الإفراز start screen collects الأرض المطلوب فرزها، نوع الإفراز، and ملاحظات then advances with التالي to attachments. Ref: §8 الإفراز — main form.",
    "Verify that attachments for الإفراز support تحميل آلي من الجهة with the same الجهة / رقم الكتاب pattern as other services. Ref: §8 — attachments.",
    # J — فك حجز
    "Verify that the applicant selects نوع فك الحجز (كلي or جزئي) and رقم القطعة then reaches the landowners grid with search (ابحث) and الإجراءات column visibility controls. Ref: §9 فك حجز — type & landowners.",
    "Verify that for جزئي only, تعديل opens the area adjustment window, saves المساحة المراد فك الحجز عنها بالمتر on حفظ التغييرات, and returns to the grid. Ref: §9 — partial unfreeze edit.",
    # K — كشف ميداني وصرف قسط
    "Verify that the applicant enters رقم القرض، المبلغ، موعد الزيارة المقترح، and ملاحظات and that عرض المصروفات السابقة lists رقم القرض، التاريخ، المبلغ، حالة القرض read-only. Ref: §10 طلب كشف ميداني وصرف قسط.",
    # L — تسوية شيكات
    "Verify that إصدار كتاب تسوية شيكات collects البنك and سبب طلب الكتاب plus إسم الساحب and ملاحظات, and that إضافة captures رقم الشيك، تاريخ الشيك، and قيمة الشيك rows. Ref: §11 إصدار كتاب تسوية شيكات.",
    # Regression / integration cross-cutting
    "Verify that when Sand, MoDEE, or CSPD retrieval fails, the UI surfaces a controlled error state and does not persist contradictory identity values. Ref: integrations — graceful failure.",
    "Verify that employee عرض can return the request to the applicant for missing data and the applicant receives a notification consistent with the return path. Ref: cross-service — return for deficiency.",
]

def main() -> None:
    base = _output_base()
    rows: list[tuple[str, str, str, str, str]] = []
    for i, desc in enumerate(_DESCS, 1):
        rows.append((f"TC-BRD-SRSFE-{i:03d}", STORY, desc, "", QA))

    base.parent.mkdir(parents=True, exist_ok=True)

    md = base.with_suffix(".md")
    md.write_text(
        "\n".join(
            [
                "# BRD-style test cases — SRS Front End ACC V1.4",
                "",
                "**Primary source (latest in repo):** `Requirements/BRD/SRS Front End - ACC – V1.4.docx`",
                "",
                "**Extract used for drafting:** `Requirements/BRD/_extracted_SRS_Front_End_ACC_V1.4.txt`",
                "",
                "**Story Title (all rows):** SRS Front End - ACC – V1.4",
                "",
                "**QA Name:** Muna Ahmad",
                "",
                "**Companion files (same test set):**",
                "",
                f"- `{md.name}` — this document",
                f"- `{base.name}.tsv` — tab-separated (5 columns)",
                f"- `{base.name}.xlsx` — Excel workbook",
                "",
                "## Assumptions",
                "",
                "- `System context & flow/SYSTEM_CONTEXT.md` and `FLOW_MAP.md` are still templates; navigation labels follow the SRS Arabic UI text.",
                "- Test environments expose Sand/MoDEE/CSPD/land integrations as configured for ACC UAT.",
                "",
                "## Open questions",
                "",
                "- Exact validation messages and field-level mandatory rules for every sub-branch of الحصول على قرض and long تمويل land-ownership variants should be refined against the controlled Word doc and UX specs.",
                "",
                "## Counts",
                "",
                f"- **Total test cases:** {len(rows)}",
                "",
                "## Sections (traceability)",
                "",
                "- **A** Shared portal — IDs TC-BRD-SRSFE-001 … 013",
                "- **B** §1 براءة ذمة — 014 … 019",
                "- **C** §2 كتاب التزام — 020 … 023",
                "- **D** §3 تغيير كفيل حسم — 024 … 029",
                "- **E** §4 كشف حساب — 030",
                "- **F** §5 قرض — 031 … 032",
                "- **G** §6 تمويل — 033 … 037",
                "- **H** §7 ضم وتوحيد — 038 … 041",
                "- **I** §8 إفراز — 042 … 043",
                "- **J** §9 فك حجز — 044 … 045",
                "- **K** §10 كشف ميداني — 046",
                "- **L** §11 تسوية شيكات — 047",
                "- **Cross** Integration / return — 048 … 049",
                "",
                "## Suggested updates to workspace knowledge",
                "",
                "- Fill `SYSTEM_CONTEXT.md` with ACC actors (مقترض، كفيل، موظف، ضامن) and integration list.",
                "- Extend `FLOW_MAP.md` with deep links or menu paths per environment once URLs are stable.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    tsv = base.with_suffix(".tsv")
    with tsv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        w.writerow(["Test Case ID", "Story Title", "Description", "Status", "QA Name"])
        w.writerows(rows)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    xlsx = base.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="DDEEFF")

    ws.append(["Test Case ID", "Story Title", "Description", "Status", "QA Name"])
    for r in rows:
        ws.append(list(r))
    for r_i in range(1, ws.max_row + 1):
        for c_i in range(1, 6):
            cell = ws.cell(r_i, c_i)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_i == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                if r_i % 2 == 0:
                    cell.fill = alt_fill
    ws.freeze_panes = "A2"
    for col, w in {"A": 20, "B": 32, "C": 110, "D": 10, "E": 14}.items():
        ws.column_dimensions[col].width = w
    wb.save(xlsx)

    print(f"Wrote {len(rows)} cases to {md.name}, {tsv.name}, {xlsx.name}")


if __name__ == "__main__":
    main()
